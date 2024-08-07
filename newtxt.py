import openpyxl as xl
import main
import osFuncts
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from student import Student


def collum_search(sheet, name):
    collum = 2
    while sheet.cell(1, collum).value.upper() != name:
        collum += 1
        if sheet.cell(1, collum).value == None:
            collum = 0
            break
    return collum


def studentList():
    students = []
    with open("alunos.txt", "r", encoding="UTF-8") as file:
        for line in file:
            aluno = line.strip().split(",")
            aluno = Student(int(aluno[0]), aluno[1])
            students.append(aluno)
    if osFuncts.file_exist_here("secundario.txt"):
        secundarios = []
        with open("secundario.txt", "r", encoding="UTF-8") as file:
            for line in file:
                aluno = line.strip().split("\t")
                secundarios.append([aluno[2], aluno[3]])
        for student in students:
            for i in range(len(secundarios)):
                if student.getName() == secundarios[i][0]:
                    student.secundario = secundarios[i][1]
                    secundarios.pop(i)
                    break
    return students


def addCabecalho(sheet, subjects):
    for i in range(1,4):
        sheet.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
    sheet.cell(1, 1).value = "Rank"
    sheet.cell(1, 2).value = "Nº"
    sheet.cell(1, 3).value = "Nome"
    cabecalho = osFuncts.getCabecalho()
    i = 0
    for year in sorted(list(cabecalho)):
        sheet.merge_cells(start_row=1, start_column=4+12*i, end_row=1, end_column=3+12*(i+1))
        sheet.cell(1, 12*i+4).value = year
        j = 0
        for semester in cabecalho[year]:
            sheet.merge_cells(start_row=2, start_column=i+4+6*j, end_row=2, end_column=i+3+6*(j+1))
            sheet.cell(2, 6*j+4+i*12).value = semester
            k=0
            for k in range(5):
                sheet.cell(3, 6*j+4+i*12+k).value = subjects[i*10+j*5+k]
            sheet.cell(3, 6*j+9+i*12).value = "Media"
            j += 1
        i += 1
    col = 4+12*i
    if osFuncts.file_exist_here("secundario.txt"):
        sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
        sheet.cell(1,col).value = "Secundario"
        sheet.column_dimensions[get_column_letter(col)].width = 10
        col += 1
    sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    sheet.cell(1,col).value = "Media"
    col += 1
    sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    sheet.cell(1,col).value = "Nº Recursos"
    sheet.column_dimensions[get_column_letter(col)].width = 12


def newTable(subjects, paths):
    students = studentList()
    for subject in subjects:
        path = paths[subject]
        if osFuncts.file_exist(f"Notas_{subject}F.xlsx", path):
            addNotetxtF(subject, path, students)
        elif osFuncts.file_exist(f"Notas_{subject}R.xlsx", path):
            addNotetxtR(subject, path, students)
        elif osFuncts.file_exist(f"Notas_{subject}.xlsx", path):
            addNotetxt(subject, path, students)

    students = sorted(students, key=lambda a:-a.getMedia())
    wb = xl.Workbook()
    sheet = wb.active
    sheet.column_dimensions['C'].width = 50
    addCabecalho(sheet, subjects)
    for i in range(len(students)):
        sheet.cell(i + 4, 1).value = i + 1
        sheet.cell(i + 4, 2).value = students[i].getNMec()
        sheet.cell(i + 4, 3).value = students[i].getName()
        saltos = 4
        total = 0
        for n in range(len(subjects)):
            note = students[i].getNote(subjects[n])
            total += note
            cell = sheet.cell(i + 4, n + saltos)
            cell.value = note
            if students[i].wasUpped(subjects[n]):
                cell.font = Font(color="0000FF")
            if note < 10:
                cell.font = Font(color="FF0000")
            if (n+1)%5==0:
                saltos += 1
                sheet.cell(i + 4, n + saltos).value = total/5
                total = 0
        n += saltos + 1
        if osFuncts.file_exist_here("secundario.txt"):
            sec = students[i].secundario.replace(",", ".")
            if sec.replace(".", "").isnumeric():
                sheet.cell(i + 4, n).value = float(sec)/10
            else:
                sheet.cell(i + 4, n).value = sec
            n += 1
        sheet.cell(i + 4, n).value = students[i].getMedia()
        n += 1
        sheet.cell(i + 4, n).value = students[i].getRecursos()
    wb.save("medias.xlsx")



def addNotetxt(subject, path, students):
    wb = xl.load_workbook(f"{path}\\Notas_{subject}.xlsx")
    sheet = wb["Table 1"]
    nota = collum_search(sheet, "NOTA")
    if nota == 0:
        print("A coluna com a nota final tem de estar indicada com NOTA")
        return
    nr = collum_search(sheet, "NR")
    alunos = []
    i = 2
    if nr == 0:
        nr = collum_search(sheet, "R")
        if nr == 0:
            print("A coluna com a nota de recurso tem de estar indicada com um NR ou R")
            return
        while sheet.cell(i, 1).value != None:
            if isinstance(sheet.cell(i, nota).value, int):
                alunos.append([sheet.cell(i, 1).value, sheet.cell(i, nota).value,0])
            else:
                alunos.append([sheet.cell(i, 1).value, 0, 0])
            if sheet.cell(i, nr).value != None:
                alunos[-1][2] += 1
            i += 1
    else:
        while sheet.cell(i, 1).value != None:
            if isinstance(sheet.cell(i, nota).value, int):
                alunos.append([sheet.cell(i, 1).value, sheet.cell(i, nota).value,0])
            else:
                alunos.append([sheet.cell(i, 1).value, 0, 0])
            notaR = sheet.cell(i, nr).value
            i += 1
            if notaR == None:
                continue
            if notaR == "F":
                continue
            alunos[-1][2] += 1
            if isinstance(notaR, str):
                continue
            if notaR > alunos[-1][1]:
                alunos[-1][1] = notaR
    updateList(students, alunos, subject)                


def addNotetxtR(subject, path, students):
    wb = xl.load_workbook(f"{path}\\Notas_{subject}.xlsx")
    sheet = wb["Table 1"]
    nota = collum_search(sheet, "NOTA")
    if nota == 0:
        print("A coluna com a nota final tem de estar indicada com NOTA")
        return
    alunos = []
    i = 2
    while sheet.cell(i, 1).value != None:
        if isinstance(sheet.cell(i, nota).value, int):
            alunos.append([sheet.cell(i, 1).value, sheet.cell(i, nota).value,0])
        else:
            alunos.append([sheet.cell(i, 1).value, 0, 0])
        i += 1
    recursoUpdate(alunos, subject, path)
    updateList(students, alunos, subject)


def addNotetxtF(subject, path, students):
    wb = xl.load_workbook(f"{path}\\Notas_{subject}.xlsx")
    sheet = wb["Table 1"]
    nota = collum_search(sheet, "NOTA")
    if nota == 0:
        print("A coluna com a nota final tem de estar indicada com NOTA")
        return
    alunos = []
    i = 2
    while sheet.cell(i, 1).value != None:
        if isinstance(sheet.cell(i, nota).value, int):
            alunos.append([sheet.cell(i, 1).value, sheet.cell(i, nota).value,0])
        else:
            alunos.append([sheet.cell(i, 1).value, 0, 0])
        i += 1
    wb = xl.load_workbook(f"{path}\\Notas_{subject}F.xlsx")
    sheet = wb["Table 1"]
    nota = collum_search(sheet, "NOTA")
    if nota == 0:
        print("A coluna com a nota final tem de estar indicada com NOTA")
        return
    i = 2
    while sheet.cell(i, 1).value != None:
        if isinstance(sheet.cell(i, nota).value, int):
            alunos.append([sheet.cell(i, 1).value, sheet.cell(i, nota).value,0])
        else:
            alunos.append([sheet.cell(i, 1).value, 0, 0])
        i += 1
    recursoUpdate(alunos, subject, path)
    updateList(students, alunos, subject)


def recursoUpdate(alunos, subject, path):
    wb = xl.load_workbook(f"{path}\\Notas_{subject}R.xlsx")
    sheet = wb["Table 1"]
    nota = collum_search(sheet, "NOTA")
    i = 2
    while sheet.cell(i, 1).value != None:
        notaR = sheet.cell(i, nota).value
        i += 1
        if notaR == None:
            continue
        if notaR == "F":
            continue
        alunoId = main.searchById(alunos, sheet.cell(i-1, 1).value)
        if alunoId == None:
            continue
        alunos[alunoId][2] += 1
        if isinstance(notaR, str):
            continue
        if notaR > alunos[alunoId][1]:
            alunos[alunoId][1] = notaR


def updateList(list, newData, subject):
    if isinstance(newData[0][0], str):
        updateListNames(list, newData, subject)
    else:
        updateListNMecs(list, newData, subject)


def updateListNMecs(list, newData, subject):
    eliminated = []
    for j in reversed(range(len(list))):
        isStudent = False
        for i in reversed(range(len(newData))):
            if list[j].getNMec() == newData[i][0]:
                list[j].addNote(subject, round(newData[i][1]))
                list[j].addRecurso(newData[i][2], subject)
                newData.pop(i)
                isStudent = True
                break
        if not isStudent:
            eliminate = main.yesno(f"O aluno {list[j][1]} não participou nesta cadeira deseja removê-lo?")
            if eliminate:
                eliminated.append(list[j][0])
                list.pop(j)
            else:
                list[j].addNote(subject, 0)
    if len(eliminated) > 0:
        removeStudent(eliminated)


def updateListNames(list, newData, subject):
    eliminated = []
    for j in reversed(range(len(list))):
        isStudent = False
        for i in reversed(range(len(newData))):
            if list[j].getName() == newData[i][0]:
                list[j].addNote(subject, round(newData[i][1]))
                list[j].addRecurso(newData[i][2], subject)
                newData.pop(i)
                isStudent = True
                break
        if not isStudent:
            eliminate = main.yesno(f"O aluno {list[j][1]} não participou nesta cadeira deseja removê-lo?")
            if eliminate:
                eliminated.append(list[j][0])
                list.pop(j)
            else:
                list[j][2].append(0)
    if len(eliminated) > 0:
        removeStudent(eliminated)


def defineAlunos(subject, path):
    wb = xl.load_workbook(path + "\\" + f"Notas_{subject}.xlsx")
    sheet = wb["Table 1"]
    alunos = []
    i = 2
    while sheet.cell(i, 1).value != None:
        alunos.append([sheet.cell(i, 1).value, sheet.cell(i, 2).value])
        i += 1
    sorted(alunos, key=lambda aluno: aluno[0])
    writeAlunos(alunos)


def writeAlunos(alunos):
    with open("alunos.txt", "w", encoding="UTF-8") as file:
        for aluno in alunos:
            for i in range(len(aluno)-1):
                file.write(str(aluno[i]) + ",")
            file.write(str(aluno[-1]) + "\n")


def removeStudent(studentList):
    students = []
    with open("alunos.txt", "r", encoding="UTF-8") as file:
        for line in file:
            aluno = line.strip().split(",")
            if int(aluno[0]) in studentList:
                continue
            students.append(aluno)
    writeAlunos(students)
